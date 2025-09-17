from flask import Flask, render_template, request, redirect, url_for, Response, flash
import cv2
from pyzbar.pyzbar import decode
import numpy as np
import os
import json
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'supersecretkey' # Thay đổi khóa bí mật này trong môi trường sản xuất
IMOU_DATA_FILE = 'imou_data.json'
HIK_DATA_FILE = 'hik_data.json'
DAHUA_DATA_FILE = 'dahua_data.json'

def read_data(data_file):
    """Đọc dữ liệu từ file JSON."""
    if not os.path.exists(data_file):
        return []
    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        return []

def write_data(data, data_file):
    """Ghi dữ liệu vào file JSON."""
    with open(data_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def parse_qr_data(qr_string):
    """Phân tích chuỗi dữ liệu từ QR code một cách chặt chẽ hơn."""
    try:
        items = qr_string.strip('{}').split(',')
        data_dict = {}
        # Yêu cầu phải có ít nhất 3 phần tử
        if len(items) < 3:
            return None

        for item in items:
            key, value = item.split(':', 1)
            data_dict[key.strip()] = value.strip()

        # Yêu cầu phải có đủ cả 3 keys SN, SC, PID
        if all(key in data_dict for key in ['SN', 'SC', 'PID']):
            return {
                'sn': data_dict.get('SN'),
                'sc': data_dict.get('SC'),
                'pid': data_dict.get('PID')
            }
        else:
            return None # Trả về None nếu thiếu bất kỳ key nào
    except Exception:
        return None

def parse_barcode_data(barcode_objects):
    """Phân tích đối tượng mã vạch đã giải mã một cách chặt chẽ hơn."""
    try:
        if barcode_objects:
            barcode_content = barcode_objects[0].data.decode('utf-8').strip()
            
            if barcode_content:
                # Thêm kiểm tra để đảm bảo đây không phải là định dạng QR của IMOU
                if any(char in barcode_content for char in ['{', '}', ':', ',']):
                    return None, 'Nội dung Barcode không đúng định dạng.'
                
                return {'sn': barcode_content}, None # Trả về dữ liệu và không có lỗi
            else:
                return None, 'Nội dung Barcode rỗng.' # Trả về None và thông báo lỗi
        else:
            return None, 'Không tìm thấy đối tượng Barcode trong ảnh.' # Trả về None và thông báo lỗi
    except Exception as e:
        return None, f'Lỗi khi phân tích Barcode: {str(e)}' # Trả về None và thông báo lỗi

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/imou', methods=['GET', 'POST'])
def imou_index():
    error = None
    temp_data = None
    if request.method == 'POST':
        if 'file' not in request.files or not request.files['file'].filename:
            error = 'Bạn chưa chọn file nào.'
        else:
            file = request.files['file']
            try:
                filestr = file.read()
                npimg = np.frombuffer(filestr, np.uint8)
                img = cv2.imdecode(npimg, cv2.IMREAD_COLOR)
                
                decoded_objects = decode(img)
                if decoded_objects:
                    qr_content = decoded_objects[0].data.decode('utf-8')
                    parsed_data = parse_qr_data(qr_content)
                    
                    if parsed_data:
                        temp_data = parsed_data # Dữ liệu tạm thời để hiển thị
                    else:
                        error = 'Mã QR không đúng định dạng chuẩn.'
                else:
                    error = 'Không tìm thấy mã QR trong ảnh.'
            except Exception as e:
                error = f'Đã xảy ra lỗi khi xử lý file: {str(e)}'

    # Lấy tham số sort từ URL
    sort_by = request.args.get('sort_by', 'stt')
    direction = request.args.get('direction', 'asc')

    records = read_data(IMOU_DATA_FILE)

    # Logic sắp xếp
    if records:
        # Mặc định sắp xếp theo chuỗi, trừ cột 'stt' sắp xếp theo số
        key_func = lambda item: int(item.get(sort_by, 0)) if sort_by == 'stt' else str(item.get(sort_by, '')).lower()
        records.sort(key=key_func, reverse=(direction == 'desc'))

    return render_template('imou.html', 
                           records=records, 
                           error=error, 
                           temp_data=temp_data, 
                           sort_by=sort_by, 
                           direction=direction)

@app.route('/imou/save', methods=['POST'])
def imou_save_data():
    """Lưu dữ liệu IMOU vào file JSON."""
    sn = request.form.get('sn')
    sc = request.form.get('sc')
    pid = request.form.get('pid')
    note = request.form.get('note', '')

    if not all([sn, sc, pid]):
        # Xử lý lỗi nếu thiếu dữ liệu, mặc dù không nên xảy ra
        return redirect(url_for('imou_index'))

    records = read_data(IMOU_DATA_FILE)
    new_record = {
        'stt': len(records) + 1,
        'sn': sn,
        'sc': sc,
        'pid': pid,
        'note': note
    }
    records.append(new_record)
    write_data(records, IMOU_DATA_FILE)
    return redirect(url_for('imou_index'))

@app.route('/imou/clear')
def imou_clear_data():
    """Hiển thị trang xác nhận xóa dữ liệu IMOU."""
    return render_template('confirm_clear.html',
                           redirect_url_on_cancel=url_for('imou_index'),
                           delete_url_on_confirm=url_for('imou_delete_confirmed'))

@app.route('/imou/delete_confirmed')
def imou_delete_confirmed():
    """Xóa dữ liệu IMOU sau khi đã xác nhận."""
    if os.path.exists(IMOU_DATA_FILE) and os.path.getsize(IMOU_DATA_FILE) > 2: # >2 để chắc chắn file không phải là "[]"
        # Tạo tên file backup với ngày giờ
        timestamp_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_filename = f"backup/imou_data_{timestamp_str}.json"
        
        # Đổi tên file data.json hiện tại thành file backup
        os.rename(IMOU_DATA_FILE, backup_filename)

    return redirect(url_for('imou_index'))

@app.route('/imou/export')
def imou_export_excel():
    """Xuất dữ liệu IMOU ra file Excel đã được định dạng."""
    records = read_data(IMOU_DATA_FILE)
    if not records:
        flash('Không có dữ liệu IMOU để xuất.', 'warning')
        return redirect(url_for('imou_index'))

    df = pd.DataFrame(records)
    # Viết hoa tiêu đề cột
    df.columns = ['STT', 'SN', 'SC', 'PID', 'NOTE']

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='ScanData')

    # Lấy các đối tượng workbook và worksheet để định dạng
    workbook = writer.book
    worksheet = writer.sheets['ScanData']

    # --- Bắt đầu định dạng ---
    # 1. Định dạng Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in worksheet[1]: # Dòng 1 là header
        cell.font = header_font
        cell.fill = header_fill

    # 2. Kẻ viền và tự động điều chỉnh độ rộng cột
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    for col_num, column_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for cell in column_cells:
            # Thêm viền cho tất cả các ô
            cell.border = thin_border
            # Tìm độ dài lớn nhất trong cột
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Điều chỉnh độ rộng cột
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    # --- Kết thúc định dạng ---

    writer.close()
    output.seek(0)

    # Tạo tên file động với ngày giờ
    timestamp_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"imou_scanned_data_{timestamp_str}.xlsx"

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

@app.route('/hik', methods=['GET', 'POST'])
def hik_index():
    error = None
    temp_data = None
    if request.method == 'POST':
        if 'file' not in request.files or not request.files['file'].filename:
            error = 'Bạn chưa chọn file nào.'
        else:
            file = request.files['file']
            try:
                filestr = file.read()
                npimg = np.frombuffer(filestr, np.uint8)
                img = cv2.imdecode(npimg, cv2.IMREAD_COLOR)
                
                decoded_objects = decode(img)
                if decoded_objects:
                    parsed_data, parse_error = parse_barcode_data(decoded_objects)
                    
                    if parsed_data:
                        temp_data = parsed_data # Dữ liệu tạm thời để hiển thị
                    else:
                        error = parse_error if parse_error else 'Mã Barcode không đúng định dạng hoặc nội dung rỗng.'
                else:
                    error = 'Không tìm thấy mã Barcode trong ảnh.'
            except Exception as e:
                error = f'Đã xảy ra lỗi khi xử lý file: {str(e)}'

    # Lấy tham số sort từ URL
    sort_by = request.args.get('sort_by', 'stt')
    direction = request.args.get('direction', 'asc')

    records = read_data(HIK_DATA_FILE)

    # Logic sắp xếp
    if records:
        # Mặc định sắp xếp theo chuỗi, trừ cột 'stt' sắp xếp theo số
        key_func = lambda item: int(item.get(sort_by, 0)) if sort_by == 'stt' else str(item.get(sort_by, '')).lower()
        records.sort(key=key_func, reverse=(direction == 'desc'))

    return render_template('hik.html', 
                           records=records, 
                           error=error, 
                           temp_data=temp_data, 
                           sort_by=sort_by, 
                           direction=direction)

@app.route('/hik/save', methods=['POST'])
def hik_save_data():
    """Lưu dữ liệu HIK vào file JSON."""
    sn = request.form.get('sn')
    note = request.form.get('note', '')

    if not sn:
        # Xử lý lỗi nếu thiếu dữ liệu, mặc dù không nên xảy ra
        return redirect(url_for('hik_index'))

    records = read_data(HIK_DATA_FILE)
    new_record = {
        'stt': len(records) + 1,
        'sn': sn,
        'note': note
    }
    records.append(new_record)
    write_data(records, HIK_DATA_FILE)
    return redirect(url_for('hik_index'))

@app.route('/hik/export')
def hik_export_excel():
    """Xuất dữ liệu HIK ra file Excel đã được định dạng."""
    records = read_data(HIK_DATA_FILE)
    if not records:
        return redirect(url_for('hik_index'))

    df = pd.DataFrame(records)
    # Viết hoa tiêu đề cột
    df.columns = ['STT', 'SN', 'NOTE']

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='ScanData')

    # Lấy các đối tượng workbook và worksheet để định dạng
    workbook = writer.book
    worksheet = writer.sheets['ScanData']

    # --- Bắt đầu định dạng ---
    # 1. Định dạng Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in worksheet[1]: # Dòng 1 là header
        cell.font = header_font
        cell.fill = header_fill

    # 2. Kẻ viền và tự động điều chỉnh độ rộng cột
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    for col_num, column_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for cell in column_cells:
            # Thêm viền cho tất cả các ô
            cell.border = thin_border
            # Tìm độ dài lớn nhất trong cột
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Điều chỉnh độ rộng cột
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    # --- Kết thúc định dạng ---

    writer.close()
    output.seek(0)

    # Tạo tên file động với ngày giờ
    timestamp_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"hik_scanned_data_{timestamp_str}.xlsx"

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

@app.route('/hik/clear')
def hik_clear_data():
    """Hiển thị trang xác nhận xóa dữ liệu HIK."""
    return render_template('confirm_clear.html',
                           redirect_url_on_cancel=url_for('hik_index'),
                           delete_url_on_confirm=url_for('hik_delete_confirmed'))

@app.route('/hik/delete_confirmed')
def hik_delete_confirmed():
    """Xóa dữ liệu HIK sau khi đã xác nhận."""
    if os.path.exists(HIK_DATA_FILE) and os.path.getsize(HIK_DATA_FILE) > 2: # >2 để chắc chắn file không phải là "[]"
        # Tạo tên file backup với ngày giờ
        timestamp_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_filename = f"backup/hik_data_{timestamp_str}.json"
        
        # Đổi tên file data.json hiện tại thành file backup
        os.rename(HIK_DATA_FILE, backup_filename)

    return redirect(url_for('hik_index'))

@app.route('/dahua', methods=['GET', 'POST'])
def dahua_index():
    error = None
    temp_data = None
    if request.method == 'POST':
        if 'file' not in request.files or not request.files['file'].filename:
            error = 'Bạn chưa chọn file nào.'
        else:
            file = request.files['file']
            try:
                filestr = file.read()
                npimg = np.frombuffer(filestr, np.uint8)
                img = cv2.imdecode(npimg, cv2.IMREAD_COLOR)
                
                decoded_objects = decode(img)
                if decoded_objects:
                    parsed_data, parse_error = parse_barcode_data(decoded_objects)
                    
                    if parsed_data:
                        temp_data = parsed_data # Dữ liệu tạm thời để hiển thị
                    else:
                        error = parse_error if parse_error else 'Mã Barcode không đúng định dạng hoặc nội dung rỗng.'
                else:
                    error = 'Không tìm thấy mã Barcode trong ảnh.'
            except Exception as e:
                error = f'Đã xảy ra lỗi khi xử lý file: {str(e)}'

    # Lấy tham số sort từ URL
    sort_by = request.args.get('sort_by', 'stt')
    direction = request.args.get('direction', 'asc')

    records = read_data(DAHUA_DATA_FILE)

    # Logic sắp xếp
    if records:
        # Mặc định sắp xếp theo chuỗi, trừ cột 'stt' sắp xếp theo số
        key_func = lambda item: int(item.get(sort_by, 0)) if sort_by == 'stt' else str(item.get(sort_by, '')).lower()
        records.sort(key=key_func, reverse=(direction == 'desc'))

    return render_template('dahua.html', 
                           records=records, 
                           error=error, 
                           temp_data=temp_data, 
                           sort_by=sort_by, 
                           direction=direction)

@app.route('/dahua/save', methods=['POST'])
def dahua_save_data():
    """Lưu dữ liệu Dahua vào file JSON."""
    sn = request.form.get('sn')
    note = request.form.get('note', '')

    if not sn:
        # Xử lý lỗi nếu thiếu dữ liệu, mặc dù không nên xảy ra
        return redirect(url_for('dahua_index'))

    records = read_data(DAHUA_DATA_FILE)
    new_record = {
        'stt': len(records) + 1,
        'sn': sn,
        'note': note
    }
    records.append(new_record)
    write_data(records, DAHUA_DATA_FILE)
    return redirect(url_for('dahua_index'))

@app.route('/dahua/clear')
def dahua_clear_data():
    """Hiển thị trang xác nhận xóa dữ liệu Dahua."""
    return render_template('confirm_clear.html',
                           redirect_url_on_cancel=url_for('dahua_index'),
                           delete_url_on_confirm=url_for('dahua_delete_confirmed'))

@app.route('/dahua/delete_confirmed')
def dahua_delete_confirmed():
    """Xóa dữ liệu Dahua sau khi đã xác nhận."""
    if os.path.exists(DAHUA_DATA_FILE) and os.path.getsize(DAHUA_DATA_FILE) > 2: # >2 để chắc chắn file không phải là "[]"
        # Tạo tên file backup với ngày giờ
        timestamp_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_filename = f"backup/dahua_data_{timestamp_str}.json"
        
        # Đổi tên file data.json hiện tại thành file backup
        os.rename(DAHUA_DATA_FILE, backup_filename)

    return redirect(url_for('dahua_index'))

@app.route('/dahua/export')
def dahua_export_excel():
    """Xuất dữ liệu Dahua ra file Excel đã được định dạng."""
    records = read_data(DAHUA_DATA_FILE)
    if not records:
        return redirect(url_for('dahua_index'))

    df = pd.DataFrame(records)
    # Viết hoa tiêu đề cột
    df.columns = ['STT', 'SN', 'NOTE']

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='ScanData')

    # Lấy các đối tượng workbook và worksheet để định dạng
    workbook = writer.book
    worksheet = writer.sheets['ScanData']

    # --- Bắt đầu định dạng ---
    # 1. Định dạng Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in worksheet[1]: # Dòng 1 là header
        cell.font = header_font
        cell.fill = header_fill

    # 2. Kẻ viền và tự động điều chỉnh độ rộng cột
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    for col_num, column_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for cell in column_cells:
            # Thêm viền cho tất cả các ô
            cell.border = thin_border
            # Tìm độ dài lớn nhất trong cột
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Điều chỉnh độ rộng cột
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    # --- Kết thúc định dạng ---

    writer.close()
    output.seek(0)

    # Tạo tên file động với ngày giờ
    timestamp_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"dahua_scanned_data_{timestamp_str}.xlsx"

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
